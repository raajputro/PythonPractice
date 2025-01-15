import os
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Alignment


# Function to plot data in to excel sheet
def write_multi_dfs_to_file(output_file, output_sheet, dataframes):
    try:
        workbook = load_workbook(output_file)
        if output_sheet in workbook.sheetnames:
            workbook.remove(workbook[output_sheet])
        worksheet = workbook.create_sheet(output_sheet)
    except FileNotFoundError:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = output_sheet

    start_row = 1
    for i in range(len(dataframes)):
        df = dataframes[i]
        start_col = 13 if (i > 0 and i % 2 == 0) else 1
        index = False if (i > 0 and i % 2 == 0) else True
        header = False if (i > 0 and i % 2 == 0) else True
        #print(f"Dataframe index: {i}, Length: {len(df)}, start row: {start_row}, start col: {start_col}")
        for r_idx, row in enumerate(dataframe_to_rows(df, index=index, header=header), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                # Apply bold font to the header row
                if r_idx == start_row and (i==0 or (i>0 and i%2!=0)):
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        start_row += len(df) + (3 if i == 0 else 0) + (3 if i % 2 == 0 else 2)
    workbook.save(output_file)

# Custom functions to slice main data to specific data
####################################################################################################################
def get_all_data(input_file, input_sheet_name):
    try:
        df = pd.read_excel(input_file, sheet_name=input_sheet_name)
        return df
    except Exception as e:
        print(f"An error occurred {e}")
        return -1

def return_sliced_dataframe(input_data_frame, filter_column, filter_value, column_names):
    try:
        df = input_data_frame
        return df[df[filter_column]==filter_value][column_names]
    except Exception as e:
        print(f"An error occurred {e}")
        return -1

def return_selected_row_count(input_df,column_name,row_value):
    try:
        count = input_df[input_df[column_name] == row_value].shape[0]
        return count
    except Exception as e:
        print(f"An error occurred {e}")
        return -1

def return_sliced_data_frame(input_df, column_name, row_value):
    try:
        return input_df[input_df[column_name]==row_value]
    except Exception as e:
        print(f"An error occurred {e}")
        return -1

def slice_month_data_ratio(month_data_ratio, month_names, attributes_name):
    sliced_month_data = {}
    for mnam in month_names:
        if mnam != 'Dec':
            sm_data = {}
            for anam in attributes_name:
                val = month_data_ratio[mnam]
                val2 = val[anam]
                sm_data[anam] = val2
            sliced_month_data[mnam] = sm_data
    test_df = pd.DataFrame(sliced_month_data)
    test_df['avg'] = round(test_df.iloc[:, 1:len(test_df.columns)].mean(axis=1))
    avg_val1 = round(float(test_df['avg'].mean()),2)
    avg_val2 = round(float(avg_val1/20),2)
    avg_df = pd.DataFrame({ 'values': [avg_val1, avg_val2] })
    return [test_df, avg_df]

####################################################################################################################
####################################################################################################################
####################################################################################################################
# Parameters
execution_directory = os.getcwd()
i_file = execution_directory + "\\Data\\Profiling Master- QC - Copy.xlsx"
i_s_name = 'Main Data'
o_file = execution_directory + "\\Data\\Profiling Master- QC - Copy.xlsx"
o_s_name = 'Scrapped_Sheet'
c_nam = ['Month','Active Listening', 'Verbal Excellence', 'Courteous Approach',
         'Identification and Action for Resolution', 'Correct & Complete Information For Resolution (CCIR)',
         'Avoid Rude/Unprofessional Behavior/Approach (ARU)','Ownership & Proctiveness (OP)']

t_df = get_all_data(input_file=i_file, input_sheet_name=i_s_name)

agent_names = t_df['Agent Name'].unique()
month_names = t_df['Month'].unique()

####################################################################################################################
for agent_name in agent_names:
#agent_name = agent_names[0]
    sliced_t_df_to_agent = return_sliced_dataframe(input_data_frame=t_df, filter_column='Agent Name',
                                                   filter_value=agent_name, column_names=c_nam)
    month_data = {}
    month_data_ratio = {}
    for month_name in month_names:
        if month_name != 'Dec':
            t_df_1 = return_sliced_data_frame(sliced_t_df_to_agent,'Month',month_name)
            td_len = len(t_df_1)
            if td_len == 0:
                td_len = 1
            month_count = {}
            month_ratio = {}
            for i in range(1, len(c_nam)):
                cname = c_nam[i]
                month_count[cname] = return_selected_row_count(input_df=t_df_1, column_name=cname, row_value='Pass')
                month_ratio[cname] = round((month_count[cname] * 100) / td_len)
            month_data[month_name] = month_count
            month_data_ratio[month_name] = month_ratio

    ################### DataFrame for Pass/Fail #######################################
    t_df_11 = pd.DataFrame(month_data)
    ################### DataFrame having ratio values ################################# to test data
    t_df_2 = pd.DataFrame(month_data_ratio)
    t_df_2['avg'] = round(t_df_2.iloc[:,1:len(t_df_2.columns)].mean(axis=1))

    ####################################################################################################################
    #creating data frames array, according to following attribute groups
    ####################################################################################################################
    com_cnames = ['Verbal Excellence', 'Avoid Rude/Unprofessional Behavior/Approach (ARU)']
    emp_cnames = ['Courteous Approach', 'Active Listening']
    bus_cnames = ['Correct & Complete Information For Resolution (CCIR)', 'Identification and Action for Resolution']
    acc_cnames = ['Ownership & Proctiveness (OP)']
    dfs_array = [t_df_11]
    smdr = slice_month_data_ratio(month_data_ratio,month_names,com_cnames)
    for x in smdr:
        dfs_array.append(x)
        #print(f"Test DF: \n {x}")
    smdr = slice_month_data_ratio(month_data_ratio,month_names,emp_cnames)
    for x in smdr:
        dfs_array.append(x)
    smdr = slice_month_data_ratio(month_data_ratio,month_names,bus_cnames)
    for x in smdr:
        dfs_array.append(x)
    smdr = slice_month_data_ratio(month_data_ratio,month_names,acc_cnames)
    for x in smdr:
        dfs_array.append(x)
    ####################################################################################################################
    # finally, write all data frames to work sheet
    ####################################################################################################################
    agent_name = agent_name[:30]
    #write_dfs_to_output_file(dataframes=dfs_array, output_file=o_file, output_sheet_name=agent_name)
    write_multi_dfs_to_file(output_file=o_file, output_sheet=agent_name, dataframes=dfs_array)
    #break
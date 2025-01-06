# importing openpyxl module
import openpyxl as xl

# source and output directory
fDirectory = "D:\\Inument_Files\\Practices"

# opening the source excel file
filename = fDirectory + "\\Excel\\SourceFile.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# opening the destination excel file
filename1 = fDirectory + "\\Excel\\OutputFile.xlsx"
wb2 = xl.load_workbook(filename1)
# ws2 = wb2.active
ws2 = wb2.worksheets[5]

# calculate total number of rows and
# columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# copying the cell values from source
# excel file to destination excel file
for i in range (1, mr + 1):
	for j in range (1, mc + 1):
		# reading cell value from source excel file
		c = ws1.cell(row = i, column = j)
		# writing the read value to destination excel file
		ws2.cell(row = i, column = j).value = c.value

# saving the destination excel file
wb2.save(str(filename1))

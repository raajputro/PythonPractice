from logging import exception
import openpyxl as xl

def returnWorkSheet (workBook, sheetName):
    if sheetName in workBook.sheetnames:
        return workBook[sheetName]

def returnColumnRowIndex (workSheet, searchText, rowOcol:bool, textPos):
    mr = workSheet.max_row
    mc = workSheet.max_column

    if rowOcol:
        for i in range (1, mr+1):
            cellPosition = workSheet.cell(row=i, column=textPos)
            cellValue = cellPosition.value
            if cellValue == searchText:
                return i
    else:
        for i in range(1, mc + 1):
            cellPosition = workSheet.cell(row=textPos, column=i)
            cellValue = cellPosition.value
            if cellValue == searchText:
                return i
    return 0

def readCellValue(ws,r,c):
    return ws.cell(row=r,column=c).value

def setCellValue(ws,r,c,value):
    ws.cell(row=r,column=c).value=value

def createPivotTable (sourceWorkSheet, destinWorkSheet):
    mc_source = sourceWorkSheet.max_column
    mr_source = sourceWorkSheet.max_row

    for i in range (1, mc_source+1):
        for j in range (1, mr_source+1):
            #cellValue_source = sourceWorkSheet.cell(row=j, column=i).value
            # destinWorkSheet.cell(row=i,column=j).value=sourceWorkSheet.cell(row=j, column=i).value
            setCellValue(destinWorkSheet,i,j,readCellValue(sourceWorkSheet,j,i))



fDirectory = "C:\\Users\\NAJIB\\Desktop\\Practices" #"D:\\Inument_Files\\Practices" # "C:\\Users\\NAJIB\\Desktop\\Practices"
sourceFile = fDirectory + "\\Excel\\SourceFile.xlsx"

try:
    wb = xl.load_workbook(sourceFile)
    ws1 = wb.worksheets[1]
    ws2 = wb.create_sheet("Pivoted10")
except Exception as e:
    print(f"An error occurred: {e}")
else:
    createPivotTable(ws1, ws2)
    wb.save(sourceFile)
finally:
    print(f"Number of worksheet: {len(wb.sheetnames)}")


#
# fDirectory = "D:\\Inument_Files\\Practices"
#
# sourceFile = fDirectory + "\\Excel\\SourceFile.xlsx"
# outputFile = fDirectory + "\\Excel\\OutputFile.xlsx"
#
# wb = xl.load_workbook(sourceFile)
# ws = wb.worksheets[0]
# sText = "ddddd"
#
# colPosition = returnColumnRowIndex(ws, sText, True, 3)
# print(sText, ' is Found in ', colPosition, ' position')
#
# wb1 = xl.load_workbook((outputFile))
# sheet_name = "Sheet5"
# ws1 = returnWorkSheet(wb1,sheet_name)
#
# if ws1:
#     print(f"The worksheet '{sheet_name} exists.")
# else:
#     print(f"The worksheet '{sheet_name} does not exist.")